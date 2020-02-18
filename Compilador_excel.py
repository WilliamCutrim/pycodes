import pandas as pd
import numpy as np
import os
import glob
import datetime
from tabula import read_pdf
from datetime import timedelta
from openpyxl import Workbook, worksheet
from openpyxl.utils.dataframe import dataframe_to_rows

Month=datetime.datetime.now().strftime("%m")
Year=datetime.datetime.now().strftime("%Y")
C_Date=Month+Year

while True:
    try:
        pasta=input('Em qual pasta estão os arquivos xls ou xlsx? ')
        indir=pasta
        os.chdir(indir)
        file_list=glob.glob("*.xls*")

        df_all=[]
        for arquivo in file_list:
            df=pd.read_excel(arquivo)
            print(df.shape, arquivo)
            df_all.append(df)
        df_all_concat=pd.concat(df_all,axis=0)

        #criando Excel
        saved_name = 'Arquivo consolidado'+C_Date+'.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = 'BASE RATEIO'
        wb.save(filename=saved_name)
        for row in dataframe_to_rows(df_all_concat, index=False, header=True):
            ws.append(row)

        for cell in ws['A'] + ws[1]:
            cell.style = 'Pandas'

        wb.save(filename=saved_name)
        wb.close
    except:
        print('Nenhum arquivo selecionado.')
